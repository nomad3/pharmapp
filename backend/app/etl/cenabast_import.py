"""
CENABAST Excel ETL Importer.

Imports two CENABAST Excel files into PostgreSQL:
  1. "Listado histórico de Ley Cenabast" — ~1,259 products with max public prices
  2. "Facturación 2020-2026 histórica Farmacias Privadas" — ~680K invoice rows

Uses openpyxl in read_only/streaming mode for memory efficiency.
Idempotent: each run DELETEs target tables before inserting.

Usage:
    python -m app.etl.cenabast_import
    python -m app.etl.cenabast_import --products /path/to/products.xlsx --invoices /path/to/invoices.xlsx
"""

import os
import sys
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.database import SessionLocal, engine
from app.models import Base
from app.models.cenabast_product import CenabastProduct
from app.models.cenabast_invoice import CenabastInvoice

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BATCH_SIZE = 1000
PROGRESS_EVERY = 10000

_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "data")

DEFAULT_PRODUCTS_FILE = os.path.join(
    _DATA_DIR, "Listado histórico de Ley Cenabast a febrero 2026.xlsx"
)
DEFAULT_ACTIVE_PMVP_FILE = os.path.join(
    _DATA_DIR, "Listado activo PMVP enero 2026.xlsx"
)
DEFAULT_INVOICES_FILE = os.path.join(
    _DATA_DIR, "Facturación 2020-2026 histórica Farmacias Privadas.xlsx"
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def safe_int(val: Any) -> Optional[int]:
    """Convert *val* to int, returning ``None`` on failure."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val: Any) -> Optional[float]:
    """Convert *val* to float, returning ``None`` on failure.

    Handles strings with comma separators (e.g. "1,388,580").
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        # Remove comma thousand-separators before converting
        cleaned = str(val).replace(",", "").strip()
        if not cleaned:
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def safe_str(val: Any) -> Optional[str]:
    """Convert *val* to a stripped string, returning ``None`` for blanks."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_fecha_doc(val: Any) -> Optional[date]:
    """Parse fecha_doc which may arrive as:
    - integer YYYYMMDD (e.g. 20250312 -> date(2025, 3, 12))
    - datetime object
    - date object
    - string in various formats
    Returns ``None`` on failure.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Try integer YYYYMMDD
    try:
        int_val = int(float(val))
        if 19000101 <= int_val <= 20991231:
            year = int_val // 10000
            month = (int_val % 10000) // 100
            day = int_val % 100
            return date(year, month, day)
    except (ValueError, TypeError):
        pass
    # Try string formats
    text_val = str(val).strip()
    if not text_val:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text_val, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _strip_accents(s: str) -> str:
    """Remove diacritical marks (accents) from a string."""
    import unicodedata
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def map_headers(sheet: Worksheet) -> Dict[str, int]:
    """Read the first row and return {normalised_header: column_index}.

    Normalisation: strip whitespace, lower-case, strip accents, collapse spaces.
    Column index is 0-based.
    """
    headers: Dict[str, int] = {}
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, cell in enumerate(row):
            if cell is not None:
                normalized = " ".join(_strip_accents(str(cell).strip().lower()).split())
                headers[normalized] = idx
    return headers


def _col(headers: Dict[str, int], *keywords: str) -> Optional[int]:
    """Return the column index whose normalised header contains any of the
    given keywords (checked in order, first match wins).  Returns ``None``
    when no match is found.
    """
    for kw in keywords:
        kw_lower = kw.lower()
        for hdr, idx in headers.items():
            if kw_lower in hdr:
                return idx
    return None


def _col_exact(headers: Dict[str, int], *keywords: str) -> Optional[int]:
    """Return the column index whose normalised header exactly matches any of
    the given keywords (checked in order, first match wins).  Returns ``None``
    when no match is found. Used when partial matching would be ambiguous.
    """
    for kw in keywords:
        kw_lower = kw.lower().strip()
        for hdr, idx in headers.items():
            if hdr == kw_lower:
                return idx
    return None


def _cell(row: tuple, col_idx: Optional[int]) -> Any:
    """Safely extract a cell value from a row tuple."""
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx]


def find_sheet(wb, keyword: str) -> Optional[Worksheet]:
    """Find a workbook sheet whose name contains *keyword* (case-insensitive)."""
    kw = keyword.lower()
    for name in wb.sheetnames:
        if kw in name.lower():
            return wb[name]
    return None


# ---------------------------------------------------------------------------
# Products importer  (Listado de PMVP)
# ---------------------------------------------------------------------------


def import_cenabast_products(db: Session, file_path: str) -> int:
    """Import the CENABAST product list (Listado de PMVP) into
    ``cenabast_products``.

    Returns the number of rows imported.
    """
    print(f"Opening products workbook: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    print(f"  Sheets found: {wb.sheetnames}")

    sheet = find_sheet(wb, "listado de pmvp")
    if sheet is None:
        # Fall back to first sheet
        sheet = find_sheet(wb, "listado")
    if sheet is None:
        print("  [SKIP] No sheet matching 'Listado de PMVP' found.")
        wb.close()
        return 0

    headers = map_headers(sheet)
    print(f"  Headers: {headers}")

    c_codigo = _col(headers, "codigo producto comercial", "codigo producto")
    c_nombre = _col(headers, "nombre producto comercial", "nombre producto")
    c_proveedor = _col(headers, "nombre proveedor", "proveedor")
    c_precio = _col(headers, "precio maximo de venta al publico", "precio maximo", "pmvp")

    # Truncate table for idempotent re-import
    db.execute(text("DELETE FROM cenabast_products"))
    db.flush()

    batch: List[CenabastProduct] = []
    seen_codes: set = set()
    count = 0

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        codigo = safe_str(_cell(row, c_codigo))
        if not codigo or codigo in seen_codes:
            continue
        seen_codes.add(codigo)

        batch.append(
            CenabastProduct(
                codigo_producto=codigo,
                nombre_producto=safe_str(_cell(row, c_nombre)),
                nombre_proveedor=safe_str(_cell(row, c_proveedor)),
                precio_maximo_publico=safe_float(_cell(row, c_precio)),
            )
        )
        count += 1

        if len(batch) >= BATCH_SIZE:
            db.bulk_save_objects(batch)
            db.flush()
            batch.clear()

        if count % PROGRESS_EVERY == 0:
            print(f"    products: {count} rows processed ...")

    if batch:
        db.bulk_save_objects(batch)
        db.flush()

    wb.close()
    print(f"  [OK] Products: {count} rows imported.")
    return count


# ---------------------------------------------------------------------------
# Active PMVP importer (Listado activo PMVP)
# ---------------------------------------------------------------------------


def import_active_pmvp(db: Session, file_path: str) -> int:
    """Import the active PMVP product list into ``cenabast_products``.

    This file has richer data (generic name, ZGEN code, contract date, active
    status). Products are UPSERTED: if a product with the same codigo_producto
    already exists (from the historical import), its extra fields are updated.

    Returns the number of rows processed.
    """
    print(f"Opening active PMVP workbook: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    print(f"  Sheets found: {wb.sheetnames}")

    sheet = find_sheet(wb, "listado productos activos")
    if sheet is None:
        sheet = find_sheet(wb, "listado")
    if sheet is None:
        print("  [SKIP] No matching sheet found.")
        wb.close()
        return 0

    headers = map_headers(sheet)
    print(f"  Headers: {headers}")

    c_zgen = _col(headers, "zgen")
    c_nombre_gen = _col(headers, "nombre genérico", "nombre generico")
    c_proveedor = _col(headers, "nombre proveedor", "proveedor")
    c_zcen = _col(headers, "zcen")
    c_descripcion = _col(headers, "descripción producto comercial", "descripcion producto")
    c_precio = _col(headers, "precio máximo venta al público", "precio maximo venta", "pmvp")
    c_fecha_cc = _col(headers, "fecha cc")
    c_activo = _col(headers, "activo")

    count = 0
    updated = 0
    inserted = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        zcen_val = safe_str(_cell(row, c_zcen))
        if not zcen_val:
            continue

        codigo = str(zcen_val)
        nombre_gen = safe_str(_cell(row, c_nombre_gen))
        proveedor = safe_str(_cell(row, c_proveedor))
        descripcion = safe_str(_cell(row, c_descripcion))
        precio = safe_float(_cell(row, c_precio))
        zgen_val = safe_str(_cell(row, c_zgen))
        fecha_cc = parse_fecha_doc(_cell(row, c_fecha_cc))
        activo_raw = safe_str(_cell(row, c_activo))
        activo = activo_raw.lower().startswith("s") if activo_raw else None

        # Try to update existing product
        existing = db.query(CenabastProduct).filter(
            CenabastProduct.codigo_producto == codigo
        ).first()

        if existing:
            existing.nombre_generico = nombre_gen
            existing.zgen = zgen_val
            existing.fecha_cc = fecha_cc
            existing.activo = activo
            if precio:
                existing.precio_maximo_publico = precio
            if proveedor:
                existing.nombre_proveedor = proveedor
            if descripcion:
                existing.nombre_producto = descripcion
            updated += 1
        else:
            db.add(CenabastProduct(
                codigo_producto=codigo,
                nombre_producto=descripcion,
                nombre_proveedor=proveedor,
                precio_maximo_publico=precio,
                zgen=zgen_val,
                nombre_generico=nombre_gen,
                fecha_cc=fecha_cc,
                activo=activo,
            ))
            inserted += 1

        count += 1
        if count % BATCH_SIZE == 0:
            db.flush()

    db.flush()
    wb.close()
    print(f"  [OK] Active PMVP: {count} processed ({updated} updated, {inserted} new).")
    return count


# ---------------------------------------------------------------------------
# Invoices importer  (Compras facturadas FP)
# ---------------------------------------------------------------------------


def import_cenabast_invoices(db: Session, file_path: str) -> int:
    """Import the CENABAST invoiced purchases (Compras facturadas FP) into
    ``cenabast_invoices``.

    This is the large dataset (~680K rows). Uses streaming mode and batch
    inserts for memory efficiency.

    Returns the number of rows imported.
    """
    print(f"Opening invoices workbook: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    print(f"  Sheets found: {wb.sheetnames}")

    sheet = find_sheet(wb, "compras facturadas")
    if sheet is None:
        sheet = find_sheet(wb, "facturadas")
    if sheet is None:
        print("  [SKIP] No sheet matching 'Compras facturadas FP' found.")
        wb.close()
        return 0

    headers = map_headers(sheet)
    print(f"  Headers: {headers}")

    # Column mapping — use exact match for ambiguous headers, partial for others
    c_fecha_doc = _col_exact(headers, "fecha doc")
    c_ano = _col_exact(headers, "ano")
    c_mes = _col_exact(headers, "mes")
    c_n_factura_sap = _col(headers, "n factura sap", "factura sap")
    c_pos = _col_exact(headers, "pos")
    c_cantidad_facturada = _col_exact(headers, "cantidad facturada")
    c_cantidad_facturada_corr = _col(headers, "cantidad facturada corregida")
    c_cantidad_unitaria = _col_exact(headers, "cantidad unitaria")
    c_cantidad_unitaria_corr = _col(headers, "cantidad unitaria corregida")
    c_division = _col_exact(headers, "division")
    c_rut_cliente_sol = _col(headers, "rut cliente solicitante")
    c_nombre_cliente_sol = _col(headers, "nombre cliente solicitante")
    c_direccion_sol = _col(headers, "direccion solicitante")
    c_comuna_sol = _col(headers, "comuna solicitante")
    c_region_sol = _col(headers, "region solicitante")
    c_rut_pagador = _col(headers, "rut pagador")
    c_cliente_dest = _col(headers, "cliente destinatario")
    c_nombre_dest = _col(headers, "nombre destinatario")
    c_direccion_dest = _col(headers, "direccion dest")
    c_comuna_dest = _col(headers, "comuna cliente dest")
    c_region_dest = _col(headers, "region cliente dest")
    # Use exact match for "valor neto" to avoid picking up "valor neto 1"
    c_valor_neto = _col_exact(headers, "valor neto")
    c_impuesto = _col_exact(headers, "impuesto")
    c_monto_bruto = _col(headers, "monto bruto")
    c_codigo_prod = _col(headers, "codigo producto comercial")
    c_nombre_prod = _col(headers, "nombre producto comercial")
    c_por = _col_exact(headers, "por")
    c_grupo_art = _col(headers, "grupo articulo")
    c_zgen = _col_exact(headers, "zgen")
    c_nombre_mat_gen = _col(headers, "nombre material generico")
    c_sector = _col_exact(headers, "sector")
    c_nombre_sector = _col(headers, "nombre sector")
    c_costo_producto = _col(headers, "costo_producto", "costo producto")
    c_margen_cenab = _col(headers, "margen_cenab", "margen cenab")
    c_margen_op_log = _col(headers, "margen_op_log", "margen op log")
    c_canal_distrib = _col(headers, "canal distrib")

    # Truncate table for idempotent re-import
    db.execute(text("DELETE FROM cenabast_invoices"))
    db.flush()

    batch: List[CenabastInvoice] = []
    count = 0

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        batch.append(
            CenabastInvoice(
                fecha_doc=parse_fecha_doc(_cell(row, c_fecha_doc)),
                ano=safe_int(_cell(row, c_ano)),
                mes=safe_int(_cell(row, c_mes)),
                n_factura_sap=safe_str(_cell(row, c_n_factura_sap)),
                pos=safe_int(_cell(row, c_pos)),
                cantidad_facturada=safe_int(_cell(row, c_cantidad_facturada)),
                cantidad_facturada_corregida=safe_int(_cell(row, c_cantidad_facturada_corr)),
                cantidad_unitaria=safe_int(_cell(row, c_cantidad_unitaria)),
                cantidad_unitaria_corregida=safe_int(_cell(row, c_cantidad_unitaria_corr)),
                division=safe_str(_cell(row, c_division)),
                rut_cliente_solicitante=safe_str(_cell(row, c_rut_cliente_sol)),
                nombre_cliente_solicitante=safe_str(_cell(row, c_nombre_cliente_sol)),
                direccion_solicitante=safe_str(_cell(row, c_direccion_sol)),
                comuna_solicitante=safe_str(_cell(row, c_comuna_sol)),
                region_solicitante=safe_str(_cell(row, c_region_sol)),
                rut_pagador=safe_str(_cell(row, c_rut_pagador)),
                cliente_destinatario=safe_str(_cell(row, c_cliente_dest)),
                nombre_destinatario=safe_str(_cell(row, c_nombre_dest)),
                direccion_dest=safe_str(_cell(row, c_direccion_dest)),
                comuna_cliente_dest=safe_str(_cell(row, c_comuna_dest)),
                region_cliente_dest=safe_str(_cell(row, c_region_dest)),
                valor_neto=safe_float(_cell(row, c_valor_neto)),
                impuesto=safe_float(_cell(row, c_impuesto)),
                monto_bruto=safe_float(_cell(row, c_monto_bruto)),
                codigo_producto_comercial=safe_str(_cell(row, c_codigo_prod)),
                nombre_producto_comercial=safe_str(_cell(row, c_nombre_prod)),
                por=safe_str(_cell(row, c_por)),
                grupo_articulo=safe_str(_cell(row, c_grupo_art)),
                zgen=safe_str(_cell(row, c_zgen)),
                nombre_material_generico=safe_str(_cell(row, c_nombre_mat_gen)),
                sector=safe_str(_cell(row, c_sector)),
                nombre_sector=safe_str(_cell(row, c_nombre_sector)),
                costo_producto=safe_float(_cell(row, c_costo_producto)),
                margen_cenab=safe_float(_cell(row, c_margen_cenab)),
                margen_op_log=safe_float(_cell(row, c_margen_op_log)),
                canal_distrib=safe_str(_cell(row, c_canal_distrib)),
            )
        )
        count += 1

        if len(batch) >= BATCH_SIZE:
            db.bulk_save_objects(batch)
            db.flush()
            batch.clear()

        if count % PROGRESS_EVERY == 0:
            print(f"    invoices: {count} rows processed ...")

    if batch:
        db.bulk_save_objects(batch)
        db.flush()

    wb.close()
    print(f"  [OK] Invoices: {count} rows imported.")
    return count


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


def import_all(
    products_file: Optional[str] = None,
    active_pmvp_file: Optional[str] = None,
    invoices_file: Optional[str] = None,
) -> None:
    """Create tables (if needed), open a DB session, and run all importers.

    The entire import runs inside a single transaction per importer.
    """
    products_path = products_file or DEFAULT_PRODUCTS_FILE
    active_pmvp_path = active_pmvp_file or DEFAULT_ACTIVE_PMVP_FILE
    invoices_path = invoices_file or DEFAULT_INVOICES_FILE

    # Ensure tables exist
    print("Creating tables if they do not exist ...")
    Base.metadata.create_all(bind=engine)

    db: Session = SessionLocal()
    try:
        n_products = 0
        n_active = 0
        n_invoices = 0

        if os.path.isfile(products_path):
            n_products = import_cenabast_products(db, products_path)
            db.commit()
        else:
            print(f"[SKIP] Products file not found: {products_path}")

        if os.path.isfile(active_pmvp_path):
            n_active = import_active_pmvp(db, active_pmvp_path)
            db.commit()
        else:
            print(f"[SKIP] Active PMVP file not found: {active_pmvp_path}")

        if os.path.isfile(invoices_path):
            n_invoices = import_cenabast_invoices(db, invoices_path)
            db.commit()
        else:
            print(f"[SKIP] Invoices file not found: {invoices_path}")

        print("\n--- Import Summary ---")
        print(f"Imported {n_products} products, {n_active} active PMVP, {n_invoices} invoices")
    except Exception:
        db.rollback()
        print("\n[ERROR] Import failed - transaction rolled back.")
        raise
    finally:
        db.close()


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Import CENABAST Excel files into PostgreSQL"
    )
    parser.add_argument(
        "--products",
        default=None,
        help=f"Path to the products Excel file (default: {DEFAULT_PRODUCTS_FILE})",
    )
    parser.add_argument(
        "--active-pmvp",
        default=None,
        help=f"Path to the active PMVP Excel file (default: {DEFAULT_ACTIVE_PMVP_FILE})",
    )
    parser.add_argument(
        "--invoices",
        default=None,
        help=f"Path to the invoices Excel file (default: {DEFAULT_INVOICES_FILE})",
    )
    args = parser.parse_args()

    import_all(
        products_file=args.products,
        active_pmvp_file=args.active_pmvp,
        invoices_file=args.invoices,
    )
