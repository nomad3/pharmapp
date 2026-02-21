"""
BMS Pharmaceutical Excel Report ETL Importer.

Imports a BMS Excel report (~32MB, 64K+ rows) into PostgreSQL using
openpyxl in read_only/streaming mode. Designed for idempotent full
re-imports: each run truncates the target BMS tables before inserting.

Usage:
    python -m app.etl.bms_import                         # default path
    python -m app.etl.bms_import /path/to/report.xlsx    # custom path
"""

import os
import sys
import unicodedata
import uuid as _uuid
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.models.bms_adjudication import BmsAdjudication
from app.models.bms_distribution import BmsDistribution
from app.models.bms_institution import BmsInstitution
from app.models.bms_purchase_order import BmsPurchaseOrder
from app.models.medication import Medication

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BATCH_SIZE = 1000
PROGRESS_EVERY = 5000

DATE_FORMATS = [
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%d.%m.%Y",
    "%Y-%m-%dT%H:%M:%S",
    "%d-%m-%Y %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _strip_accents(s: str) -> str:
    """Remove diacritical marks (accents) from a string."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def find_sheet(wb, keyword: str) -> Optional[Worksheet]:
    """Find a workbook sheet whose name contains *keyword* (case-insensitive)."""
    kw = _strip_accents(keyword.lower())
    for name in wb.sheetnames:
        if kw in _strip_accents(name.lower()):
            return wb[name]
    return None


def map_headers(sheet: Worksheet, header_row: int = 1) -> Dict[str, int]:
    """Read *header_row* (1-based) and return {normalised_header: column_index}.

    Normalisation: strip whitespace, lower-case, collapse multiple spaces,
    remove accents, replace underscores with spaces.
    Column index is 0-based.
    """
    headers: Dict[str, int] = {}
    for row in sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        for idx, cell in enumerate(row):
            if cell is not None:
                normalized = " ".join(str(cell).strip().lower().split())
                normalized = _strip_accents(normalized)
                normalized = normalized.replace("_", " ")
                headers[normalized] = idx
    return headers


def _col(headers: Dict[str, int], *keywords: str) -> Optional[int]:
    """Return the column index whose normalised header contains any of the
    given keywords (checked in order, first match wins).  Returns ``None``
    when no match is found.
    """
    for kw in keywords:
        kw_lower = _strip_accents(kw.lower().replace("_", " "))
        for hdr, idx in headers.items():
            if kw_lower in hdr:
                return idx
    return None


def _cell(row: tuple, col_idx: Optional[int]) -> Any:
    """Safely extract a cell value from a row tuple."""
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx]


def parse_date(val: Any) -> Optional[date]:
    """Convert a value to a ``datetime.date``.

    Handles native datetime objects, date objects, and a variety of string
    formats commonly seen in Chilean / BMS reports.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text_val = str(val).strip()
    if not text_val:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text_val, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def safe_int(val: Any) -> Optional[int]:
    """Convert *val* to int, returning ``None`` on failure."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val: Any) -> Optional[float]:
    """Convert *val* to float, returning ``None`` on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_str(val: Any) -> Optional[str]:
    """Convert *val* to a stripped string, returning ``None`` for blanks."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Medication lookup cache
# ---------------------------------------------------------------------------


def _build_medication_lookup(db: Session) -> Dict[tuple, str]:
    """Return a dict mapping ``(active_ingredient_lower, dosage_lower)``
    to the medication UUID (as string) for FK matching.
    """
    lookup: Dict[tuple, str] = {}
    for med in db.query(Medication).all():
        ai = (med.active_ingredient or "").strip().lower()
        dosage = (med.dosage or "").strip().lower()
        if ai:
            lookup[(ai, dosage)] = str(med.id)
    return lookup


# ---------------------------------------------------------------------------
# Per-sheet importers
# ---------------------------------------------------------------------------


def import_institutions(db: Session, wb) -> int:
    """Import *Maestro Instituciones* sheet into ``bms_institutions``.

    This sheet has a quirk: row 1 is data, row 2 has headers, data resumes row 3.
    We detect this by trying row 1 then row 2 for recognizable headers.
    """
    sheet = find_sheet(wb, "instituciones")
    if sheet is None:
        print("  [SKIP] No sheet matching 'instituciones' found.")
        return 0

    # Try row 1 first, then row 2 to find headers
    headers = map_headers(sheet, header_row=1)
    data_start = 2
    if not any("rut" in h for h in headers):
        headers = map_headers(sheet, header_row=2)
        data_start = 3

    # Use exact match for "cliente" to avoid matching "rut cliente"
    c_rut = headers.get("rut cliente") or _col(headers, "rut")
    c_code = headers.get("cliente") or _col(headers, "codigo cliente", "client code")
    c_razon = _col(headers, "razon social")
    c_region = _col(headers, "region")
    c_comuna = _col(headers, "comuna")

    print(f"    headers detected at row {data_start - 1}: {list(headers.keys())[:5]}")

    db.execute(text("DELETE FROM bms_institutions"))
    db.flush()

    batch: List[BmsInstitution] = []
    count = 0
    seen_ruts = set()

    for row_num, row in enumerate(sheet.iter_rows(min_row=data_start, values_only=True), start=data_start):
        rut_val = _safe_str(_cell(row, c_rut))
        if not rut_val or rut_val in seen_ruts:
            continue
        seen_ruts.add(rut_val)

        batch.append(
            BmsInstitution(
                rut=rut_val,
                client_code=safe_int(_cell(row, c_code)),
                razon_social=_safe_str(_cell(row, c_razon)),
                region=_safe_str(_cell(row, c_region)),
                comuna=_safe_str(_cell(row, c_comuna)),
            )
        )
        count += 1

        if len(batch) >= BATCH_SIZE:
            db.bulk_save_objects(batch)
            db.flush()
            batch.clear()

        if count % PROGRESS_EVERY == 0:
            print(f"    institutions: {count} rows processed …")

    if batch:
        db.bulk_save_objects(batch)
        db.flush()

    print(f"  [OK] Institutions: {count} rows imported.")
    return count


def import_distribution(db: Session, wb, med_lookup: Dict[tuple, str]) -> int:
    """Import *Data Distribución* sheet into ``bms_distributions``."""
    sheet = find_sheet(wb, "data distribución") or find_sheet(wb, "data distribucion") or find_sheet(wb, "distribución")
    if sheet is None:
        print("  [SKIP] No sheet matching 'distribución' found.")
        return 0
    print(f"    Using sheet: {sheet.title}")

    headers = map_headers(sheet)

    c_ai = _col(headers, "p activo pht", "principio activo")
    c_comp = _col(headers, "comp pht", "composicion")
    c_meas = _col(headers, "medida pht", "medida")
    c_pcom = _col(headers, "nombre producto comercial", "producto comercial")
    c_pgen = _col(headers, "nombre producto generico", "producto generico")
    c_rut = _col(headers, "rut destinatario", "rut institucion", "rut")
    c_dest = _col(headers, "nombre cliente destinatario", "cliente destinatario", "cliente destino")
    c_reg = _col(headers, "nombre region", "region")
    c_com = _col(headers, "comuna")
    c_ss = _col(headers, "servicio de salud", "servicio salud")
    c_date = _col(headers, "fecha de entrega", "fecha entrega", "fecha despacho")
    c_po = _col(headers, "orden de compra")
    c_sd = _col(headers, "documento de venta")
    c_oq = _col(headers, "cantidad de pedido", "cantidad oc")
    c_uq = _col(headers, "cantidad unitaria")
    c_upp = _col(headers, "cantidad por envase", "unidades por envase")
    c_ga = headers.get("monto bruto") or _col(headers, "monto bruto")
    c_na = _col(headers, "monto neto")
    c_gup = _col(headers, "precio unitario bruto")
    c_nup = _col(headers, "precio unitario neto")
    c_mkt = _col(headers, "market", "mercado")
    c_bms = _col(headers, "bms/competencia", "bms")
    c_prov = _col(headers, "nombre proveedor", "proveedor")
    c_chan = _col(headers, "nombre canal de distribucion", "canal distribucion")

    db.execute(text("DELETE FROM bms_distributions"))
    db.flush()

    batch: List[BmsDistribution] = []
    count = 0

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Attempt medication FK lookup
        ai_val = _safe_str(_cell(row, c_ai))
        comp_val = _safe_str(_cell(row, c_comp))
        med_id = None
        if ai_val:
            key = (ai_val.lower(), (comp_val or "").lower())
            med_id_str = med_lookup.get(key)
            if med_id_str is None:
                # Try matching on active_ingredient alone (dosage empty)
                med_id_str = med_lookup.get((ai_val.lower(), ""))
            if med_id_str:
                med_id = _uuid.UUID(med_id_str)

        batch.append(
            BmsDistribution(
                active_ingredient=ai_val,
                composition=comp_val,
                measure=_safe_str(_cell(row, c_meas)),
                product_commercial_name=_safe_str(_cell(row, c_pcom)),
                product_generic_name=_safe_str(_cell(row, c_pgen)),
                medication_id=med_id,
                institution_rut=_safe_str(_cell(row, c_rut)),
                client_destination_name=_safe_str(_cell(row, c_dest)),
                region=_safe_str(_cell(row, c_reg)),
                comuna=_safe_str(_cell(row, c_com)),
                servicio_salud=_safe_str(_cell(row, c_ss)),
                delivery_date=parse_date(_cell(row, c_date)),
                purchase_order=_safe_str(_cell(row, c_po)),
                sale_document=_safe_str(_cell(row, c_sd)),
                order_quantity=safe_int(_cell(row, c_oq)),
                unit_quantity=safe_int(_cell(row, c_uq)),
                units_per_package=safe_int(_cell(row, c_upp)),
                gross_amount=safe_float(_cell(row, c_ga)),
                net_amount=safe_float(_cell(row, c_na)),
                gross_unit_price=safe_float(_cell(row, c_gup)),
                net_unit_price=safe_float(_cell(row, c_nup)),
                market=_safe_str(_cell(row, c_mkt)),
                bms_competition=_safe_str(_cell(row, c_bms)),
                provider_name=_safe_str(_cell(row, c_prov)),
                distribution_channel=_safe_str(_cell(row, c_chan)),
            )
        )
        count += 1

        if len(batch) >= BATCH_SIZE:
            db.bulk_save_objects(batch)
            db.flush()
            batch.clear()

        if count % PROGRESS_EVERY == 0:
            print(f"    distribution: {count} rows processed …")

    if batch:
        db.bulk_save_objects(batch)
        db.flush()

    print(f"  [OK] Distribution: {count} rows imported.")
    return count


def import_purchase_orders(db: Session, wb) -> int:
    """Import *Data Orden de Compra* sheet into ``bms_purchase_orders``."""
    sheet = find_sheet(wb, "orden de compra")
    if sheet is None:
        print("  [SKIP] No sheet matching 'orden de compra' found.")
        return 0

    headers = map_headers(sheet)

    c_pa = _col(headers, "pactivo", "p.activo", "principio activo")
    c_pres = _col(headers, "presentacion")
    c_med = _col(headers, "un medida pht", "medida")
    c_cpht = _col(headers, "cant pht", "cantidad pht")
    c_ppht = _col(headers, "precio pht")
    c_vt = _col(headers, "valor total")
    c_fecha = headers.get("fecha") or _col(headers, "fecha")
    c_inst = _col(headers, "instituciones", "institucion")
    c_reg = headers.get("region") or _col(headers, "region")
    c_com = _col(headers, "comuna cliente", "comuna")
    c_sup = _col(headers, "sucursal proveedor", "supplier")
    c_corp = _col(headers, "corporacionespht", "corporacion")
    c_mkt = _col(headers, "market", "mercado")
    c_bms = _col(headers, "bms/competencia", "bms")
    c_tipo = _col(headers, "tipo oc")
    c_lic = _col(headers, "id licitacion")

    db.execute(text("DELETE FROM bms_purchase_orders"))
    db.flush()

    batch: List[BmsPurchaseOrder] = []
    count = 0

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        batch.append(
            BmsPurchaseOrder(
                pactivo=_safe_str(_cell(row, c_pa)),
                presentacion=_safe_str(_cell(row, c_pres)),
                medida=_safe_str(_cell(row, c_med)),
                cant_pht=safe_int(_cell(row, c_cpht)),
                precio_pht=safe_float(_cell(row, c_ppht)),
                valor_total=safe_float(_cell(row, c_vt)),
                fecha=parse_date(_cell(row, c_fecha)),
                institucion=_safe_str(_cell(row, c_inst)),
                region=_safe_str(_cell(row, c_reg)),
                comuna=_safe_str(_cell(row, c_com)),
                supplier=_safe_str(_cell(row, c_sup)),
                corporation=_safe_str(_cell(row, c_corp)),
                market=_safe_str(_cell(row, c_mkt)),
                bms_competition=_safe_str(_cell(row, c_bms)),
                tipo_oc=_safe_str(_cell(row, c_tipo)),
                id_licitacion=_safe_str(_cell(row, c_lic)),
            )
        )
        count += 1

        if len(batch) >= BATCH_SIZE:
            db.bulk_save_objects(batch)
            db.flush()
            batch.clear()

        if count % PROGRESS_EVERY == 0:
            print(f"    purchase_orders: {count} rows processed …")

    if batch:
        db.bulk_save_objects(batch)
        db.flush()

    print(f"  [OK] Purchase orders: {count} rows imported.")
    return count


def import_adjudications(db: Session, wb) -> int:
    """Import *Data Adjudicaciones* sheet into ``bms_adjudications``."""
    sheet = find_sheet(wb, "data adjudicaciones") or find_sheet(wb, "adjudicaciones")
    if sheet is None:
        print("  [SKIP] No sheet matching 'adjudicaciones' found.")
        return 0
    print(f"    Using sheet: {sheet.title}")

    headers = map_headers(sheet)

    c_adq = _col(headers, "adquisicion")
    c_rut = _col(headers, "rut cliente")
    c_fecha = _col(headers, "fechaadjudicacion", "fecha adjudicacion")
    c_est = _col(headers, "estado")
    c_pa = _col(headers, "pactivo", "p.activo", "principio activo")
    c_comp = _col(headers, "composicion")
    c_pres = _col(headers, "presentacion")
    c_pu = _col(headers, "precio unit", "precio unitario")
    c_ca = _col(headers, "cantadjudicada", "cantidad adjudicada")
    c_va = _col(headers, "valor adjudicado")
    c_rs = _col(headers, "razonsocialcliente", "razon social cliente")
    c_cp = _col(headers, "corpproveedor", "corp proveedor")
    c_mkt = _col(headers, "market", "mercado")
    c_bms = _col(headers, "bms/competencia", "bms")

    db.execute(text("DELETE FROM bms_adjudications"))
    db.flush()

    batch: List[BmsAdjudication] = []
    count = 0

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        batch.append(
            BmsAdjudication(
                adquisicion=_safe_str(_cell(row, c_adq)),
                rut_cliente=_safe_str(_cell(row, c_rut)),
                fecha_adjudicacion=parse_date(_cell(row, c_fecha)),
                estado=_safe_str(_cell(row, c_est)),
                pactivo=_safe_str(_cell(row, c_pa)),
                composicion=_safe_str(_cell(row, c_comp)),
                presentacion=_safe_str(_cell(row, c_pres)),
                precio_unit=safe_float(_cell(row, c_pu)),
                cant_adjudicada=safe_int(_cell(row, c_ca)),
                valor_adjudicado=safe_float(_cell(row, c_va)),
                razon_social_cliente=_safe_str(_cell(row, c_rs)),
                corp_proveedor=_safe_str(_cell(row, c_cp)),
                market=_safe_str(_cell(row, c_mkt)),
                bms_competition=_safe_str(_cell(row, c_bms)),
            )
        )
        count += 1

        if len(batch) >= BATCH_SIZE:
            db.bulk_save_objects(batch)
            db.flush()
            batch.clear()

        if count % PROGRESS_EVERY == 0:
            print(f"    adjudications: {count} rows processed …")

    if batch:
        db.bulk_save_objects(batch)
        db.flush()

    print(f"  [OK] Adjudications: {count} rows imported.")
    return count


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


def import_all(file_path: str) -> None:
    """Open the BMS Excel workbook, connect to the database, and import all
    sheets in order.  The entire import runs inside a single transaction.
    """
    print(f"Opening workbook: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    print(f"  Sheets found: {wb.sheetnames}")

    db: Session = SessionLocal()
    try:
        # Build medication FK lookup
        print("Building medication lookup cache …")
        med_lookup = _build_medication_lookup(db)
        print(f"  {len(med_lookup)} medication entries cached.")

        # Import each sheet
        n_inst = import_institutions(db, wb)
        n_dist = import_distribution(db, wb, med_lookup)
        n_po = import_purchase_orders(db, wb)
        n_adj = import_adjudications(db, wb)

        db.commit()
        print("\n--- Import Summary ---")
        print(
            f"Imported {n_inst} institutions, {n_dist} distribution records, "
            f"{n_po} purchase orders, {n_adj} adjudications"
        )
    except Exception:
        db.rollback()
        print("\n[ERROR] Import failed — transaction rolled back.")
        raise
    finally:
        db.close()
        wb.close()


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        # Default: backend/data/bms_report.xlsx relative to this file's
        # location (app/etl/bms_import.py → ../../data/bms_report.xlsx)
        here = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(here, os.pardir, os.pardir, "data", "bms_report.xlsx")
        path = os.path.normpath(path)

    if not os.path.isfile(path):
        print(f"File not found: {path}")
        sys.exit(1)

    import_all(path)
